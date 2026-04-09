import csv
import io
import json
from datetime import datetime

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import AssetColumn


class AssetExportService:
    STANDARD_COLUMNS = [
        ("asset_number", "Asset Number"),
        ("hostname", "Hostname"),
        ("ip_address", "IP Address"),
        ("asset_type", "Asset Type"),
        ("asset_level", "Asset Level"),
        ("description", "Description"),
        ("is_alive", "Is Alive"),
        ("created_at", "Created At"),
        ("updated_at", "Updated At"),
        ("created_by", "Created By"),
        ("updated_by", "Updated By"),
    ]

    def __init__(self, queryset):
        self.assets = list(queryset)
        self.custom_columns = self._build_custom_columns()

    def export(self, export_format="xlsx"):
        if export_format == "csv":
            return self.export_csv()
        return self.export_xlsx()

    def get_filename(self, export_format="xlsx"):
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        return f"cmdb_assets_{timestamp}.{export_format}"

    def get_content_type(self, export_format="xlsx"):
        if export_format == "csv":
            return "text/csv; charset=utf-8"
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def _build_custom_columns(self):
        configured_columns = list(AssetColumn.objects.order_by("id").values_list("name", "label"))
        configured_names = {name for name, _ in configured_columns}
        discovered_names = set()

        for asset in self.assets:
            if isinstance(asset.custom_attributes, dict):
                discovered_names.update(asset.custom_attributes.keys())

        extra_columns = [(name, name) for name in sorted(discovered_names - configured_names)]
        return configured_columns + extra_columns

    def _headers(self):
        return [label for _, label in self.STANDARD_COLUMNS] + [label for _, label in self.custom_columns]

    def _rows(self):
        for asset in self.assets:
            row = []
            for field_name, _ in self.STANDARD_COLUMNS:
                row.append(self._resolve_standard_field(asset, field_name))

            custom_attributes = asset.custom_attributes if isinstance(asset.custom_attributes, dict) else {}
            for name, _ in self.custom_columns:
                row.append(self._serialize_value(custom_attributes.get(name, "")))
            yield row

    def _resolve_standard_field(self, asset, field_name):
        if field_name == "created_by":
            return getattr(asset.created_by, "username", "")
        if field_name == "updated_by":
            return getattr(asset.updated_by, "username", "")
        return self._serialize_value(getattr(asset, field_name, ""))

    def _serialize_value(self, value):
        if value is None:
            return ""
        if isinstance(value, bool):
            return "Yes" if value else "No"
        if isinstance(value, datetime):
            return timezone.localtime(value).strftime("%Y-%m-%d %H:%M:%S") if timezone.is_aware(value) else value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)

    def export_csv(self):
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(self._headers())
        for row in self._rows():
            writer.writerow(row)
        return buffer.getvalue().encode("utf-8-sig")

    def export_xlsx(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "CMDB Assets"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
        alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        header_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        headers = self._headers()
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for row in self._rows():
            worksheet.append(row)

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = alignment
                cell.border = border

        worksheet.freeze_panes = "A2"

        for index, column_cells in enumerate(worksheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 12), 40)

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

